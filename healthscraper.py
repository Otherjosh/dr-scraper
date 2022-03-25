# -*- coding: utf-8 -*-
"""
Created on Sat Dec 11 16:55:00 2021

@author: joshu
"""

import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
import undetected_chromedriver.v2 as uc 
import openpyxl 
import os
from datetime import datetime
from selenium.webdriver.common.by import By
import re
import pandas as pd
from selenium.common.exceptions import NoSuchElementException

    
def clearSurvey(driver):
    #sometimes there is a survey that pops up, this dismisses that 
    noThanks = driver.find_element(By.XPATH, '//button[@class="_hj-wTnOw__SurveyInvitation__noThanksButton _hj-3OscV__styles__clearButton"]')
    noThanks.click()

# driver.find_element(By.XPATH'//h3[classclass="zero-results__title"]')    


def countDupes(file):      
    #this uses pandas to find the # of repeat rows for each doctor to pass into checkDupes fx 
    df = pd.ExcelFile(file).parse('Medicare_Part_D_Prescribers_by_') #you could add index_col=0 if there's an index
    dfNPI = df['Prscrbr_NPI']
    dupes = dfNPI.groupby((dfNPI != dfNPI.shift()).cumsum()).cumcount()
    dupesDict = {}
    rows = df.shape[0]
    
    for i in range(1, rows):
        print(f"NPI: {dfNPI[i]}, dupes: {dupes[i]}")
        if dupes[i-1] > dupes[i]:
            prevNPI = dfNPI[i-1]
            prevDupes = dupes[i-1]
            dupesDict[prevNPI] = prevDupes
    #print(dupesDict)
    return dupesDict
    #df.to_excel() could write back to excel
    
    
def checkDupes(currentDoctor, sheet_obj, i, Prscrbr_NPI):
    #if the phone num is being stored on the first row for a DR correctly, we can just pull that and use it to assign for the rest of the dupes 
    nextNPI = sheet_obj.cell(row = i+1, column = 1).value
    print("next NPI: " + str(nextNPI))
    #print("newRow in checkDupes: " + str(newRow))
    address = currentDoctor.clinicAddress
    name = currentDoctor.clinicName
    phone = currentDoctor.clinicPhone
    
    if nextNPI == currentDoctor.NPI:  #this should fix the blank row for first dupe
        print(f"writing dupe info \naddress: {address} \nname: {name} \nphone: {phone}")
        addressCell = sheet_obj.cell(row=i, column=6)
        addressCell.value = address
        nameCell = sheet_obj.cell(row=i, column=7)
        nameCell.value = name
        phoneCell = sheet_obj.cell(row=i, column=8)
        phoneCell.value = phone
        
        addressCell = sheet_obj.cell(row=i+1, column=6)
        addressCell.value = address
        nameCell = sheet_obj.cell(row=i+1, column=7)
        nameCell.value = name
        phoneCell = sheet_obj.cell(row=i+1, column=8)
        phoneCell.value = phone
        
        i = i+1  #its writing current row + the next one, so I think this should be 1

    else:
        i+=1

    print("end of check dupes, i = " + str(i))
    return i
    #return b
    #infinite loop here for dupes that aren't found in the search results 
    
    
    #maybe iterate forward with i here and then return i? 
    #pass 
    
    
    
def findDoctor(driver, currentDoctor):
    #this function finds the currentDoctor on healthgrades 
    driver.get('https://www.healthgrades.com')
    driver.maximize_window()
    skip = False
    try:
        clearSurvey(driver)
    except:
        pass
    time.sleep(2)
    try:  #sometimes there is a satisfaction survey, this clears that if present
        dismissSatisfaction = driver.find_element(By.XPATH, '//button[@class="_hj-OO1S1__styles__openStateToggle"]')
        dismissSatisfaction.click()
    except:
        pass
    drName = currentDoctor.firstName + ' ' + currentDoctor.lastName
    nameBox = driver.find_element(By.XPATH, '//input[@name="term-input-group"]')
    # nameBox.click()
    # nameBox.send_keys(Keys.CONTROL + "A")
    nameBox.send_keys(drName)  #send Prscrbr_Last_Org_Name + ' ' + Prscrbr_First_Name here
    print("entered name: " + drName)
    time.sleep(1)
    drLocation = currentDoctor.city + ', ' + currentDoctor.state
    locationBox = driver.find_element(By.XPATH, '//input[@name="location-input-group"]')
    locationBox.click()
    locationBox.send_keys(Keys.CONTROL + "A")
    time.sleep(1)
    
    locationBox.send_keys(drLocation)  #send Prscrbr_City + ', ' +	Prscrbr_State_Abrvtn here
    print("entered location: " + drLocation)
    time.sleep(1.5)
    searchButton = driver.find_element(By.XPATH, '//button[@class="search-bar-btns__search-btn"]')
    searchButton.click() 
    time.sleep(5)
    
    addressList = driver.find_elements(By.XPATH, '//div[@class="location-info__office-loc"]/div')   
    for item in addressList:
        print(item.get_attribute('innerHTML'))
    try:  #this tried to dismiss a survey, does nothing if there isn't one present 
        dismissSurvey = driver.find_element(By.XPATH, '//button[@class="_hj-OO1S1__styles__openStateToggle"]')
        dismissSurvey.click()
    except:
        pass

    try:
        targetResult = driver.find_element(By.XPATH, '//a[@class="provider-name__lnk"]')  #this should get the first result, possibly revisit this if searches are providing multiple results or first result isn't correct 
        targetResult.click()  #this will open the details for the DR in question 
    except:
        skip = True
        print("unable to find target result, skipping")
        print(driver.title)
        print("not switching to window[1] due to skip == true")
    time.sleep(1)
    if skip == False:
        driver.switch_to.window(driver.window_handles[1])
    print(driver.title)
    return skip
    

def getData(driver, currentDoctor):
    #this will gather the data from the details page 
    clinicName = ''
    clinicAddress = ''
    clinicPhone = ''
    try:
        cookiesPopup = driver.find_element(By.XPATH, '//button[@class="onetrust-close-btn-handler banner-close-button ot-close-icon"]')
        cookiesPopup.click()
    except:
        pass
    try:
        clearSurvey(driver)
    except: 
        pass
    time.sleep(5)
    print("inside getData, tab title: " + driver.title)
    
  
    try:  #this tries to get the data from the location section on the result page 
        addressLIST = driver.find_elements(By.XPATH, '//div[@class="office-location profile-subsection-bordered-container"]/section/div/address/div/span')
        addressUnedited = ''
        for item in addressLIST:
            addressUnedited = addressUnedited + item.get_attribute('innerHTML')
        clinicAddress = addressUnedited.replace('<!-- -->', '')  
        
        clinicName = driver.find_element(By.XPATH, '//div[@class="office-location profile-subsection-bordered-container"]/section/div/address/div/a').get_attribute('innerHTML')
        clinicPhone = driver.find_element(By.XPATH, '//div[@class="office-location profile-subsection-bordered-container"]/section/div[2]/div/a').get_attribute('innerHTML')
        
        currentDoctor.add_clinic_address(clinicAddress)
        print("added clinic address: " + currentDoctor.clinicAddress)
        currentDoctor.add_clinic_name(clinicName)
        print("added clinic name: " + currentDoctor.clinicName)
        currentDoctor.add_clinic_phone(clinicPhone)
        print("added clinic phone: " + currentDoctor.clinicPhone)
        return currentDoctor
    except: 
        pass 
    
     
        
    try: #make this a try
        #this gets the clinic name & address by splitting up <address> element and it's children 
        print("in address plan B")
        address = driver.find_element(By.XPATH, '//address[@class="address"]')
        addressChildren = address.find_elements(By.XPATH, './/*')
        holder = ''
        for element in addressChildren:
            if len(element.get_attribute('innerHTML')) > 1:
                holder = holder + element.get_attribute('innerHTML')
                #print(holder)
        regex = r"<(.+?)>"
        result = re.sub(regex, 'zzzzz', holder, 0, re.MULTILINE)
        #print(result)
        resultSplit = result.split('zzzzz')
        clinicName = resultSplit[1]
        print(clinicName)
        clinicAddress = resultSplit[3] + ' ' + resultSplit[6] + ', ' + resultSplit[9] + ' ' + resultSplit[12]
        print(clinicAddress)
        
        try:
            clinicPhone = driver.find_element(By.XPATH, '//div[@class="office-location profile-subsection-bordered-container"]/section/div[2]/div/a').get_attribute('innerHTML')
        except: 
            pass
        
        if len(clinicPhone) < 5:
            try: #tries the phone number by itself, as its possible for clinic/name address to not match the original technique but the phone number still does 
                clinicPhone = driver.find_element(By.XPATH, '//a[@class="toggle-phone-number-button"]').get_attribute('innerHTML')
            except: #skips the phone number if it still isn't able to get it 
                pass
        
        currentDoctor.add_clinic_address(clinicAddress)
        print("added clinic address: " + currentDoctor.clinicAddress)
        currentDoctor.add_clinic_name(clinicName)
        print("added clinic name: " + currentDoctor.clinicName)
        currentDoctor.add_clinic_phone(clinicPhone)
        print("added clinic phone: " + currentDoctor.clinicPhone)
        return currentDoctor
        
        print("plan b fail")
        
        
    except NoSuchElementException: 
        #maybe log this 
        print("this should happen after plan b")
        pass  
        
    if len(clinicPhone) < 5:
        try: #tries the phone number by itself, as its possible for clinic/name address to not match the original technique but the phone number still does 
            clinicPhone = driver.find_element(By.XPATH, '//a[@class="toggle-phone-number-button"]').get_attribute('innerHTML')
        except: #skips the phone number if it still isn't able to get it 
            pass
        
        
    
    currentDoctor.add_clinic_address(clinicAddress)
    print("added clinic address: " + currentDoctor.clinicAddress)
    currentDoctor.add_clinic_name(clinicName)
    print("added clinic name: " + currentDoctor.clinicName)
    currentDoctor.add_clinic_phone(clinicPhone)
    print("added clinic phone: " + currentDoctor.clinicPhone)
    return currentDoctor
    

def logData(i, currentDoctor, sheet_obj, driver, dupesDict):
    #this will write clinic address, clinic name, and clinic phone number to the excel sheet in columns 6, 7, 8
    #column numbers start at 1 not 0 
    
    currentNPI = sheet_obj.cell(row=i, column=1).value
    nextNPI = sheet_obj.cell(row = i+1, column = 1).value
    
    addressCell = sheet_obj.cell(row=i, column=6)
    addressCell.value = currentDoctor.clinicAddress
    nameCell = sheet_obj.cell(row=i, column=7)
    nameCell.value = currentDoctor.clinicName
    phoneCell = sheet_obj.cell(row=i, column=8)
    phoneCell.value = currentDoctor.clinicPhone 
#&amp; in names can be .replaced to & maybe    
    print("logged data here: ")
    print(currentDoctor.clinicAddress)
    print(currentDoctor.clinicName)
    print(currentDoctor.clinicPhone)
    try:
        dupes = dupesDict[currentNPI]
        newRow = i + dupes
        while currentDoctor.NPI == nextNPI and i < newRow:
            print("in logData while loop try, i = " + str(i))
            i = checkDupes(currentDoctor, sheet_obj, i, currentNPI)
            #i+=1
    except:
        while currentDoctor.NPI == nextNPI:
            print("in logData while loop except, i = " + str(i))
            i = checkDupes(currentDoctor, sheet_obj, i, currentNPI)
        
        
    
    #driver.find_element(By.XPATH, '//div[@class="inline-contact-container"]/div/div/p').send_keys(Keys.COMMAND + 'W')
    if driver.current_window_handle != driver.window_handles[0]:
        driver.close()
    driver.switch_to.window(driver.window_handles[0])
    time.sleep(1)
    homeButton = driver.find_element(By.XPATH, '//a[@class="hgGlobalHeader__Logo"]')
    homeButton.click()
    print("active window at end of logData: " + driver.title)
    return i
    

def readExcel(currentDoctor, sheet_obj, i, driver, b):
    #This function reads the data from the excel sheet. column and row numbers start at 1, not 0 
    Prscrbr_NPI = sheet_obj.cell(row = i, column = 1).value
    Prscrbr_Last_Org_Name = sheet_obj.cell(row = i, column = 2).value
    Prscrbr_First_Name = sheet_obj.cell(row = i, column = 3).value
    Prscrbr_City = sheet_obj.cell(row = i, column = 4).value
    Prscrbr_State_Abrvtn = sheet_obj.cell(row = i, column = 5).value
    
    # nextNPIcell = sheet_obj.cell(row = i+1, column = 1)
    # print(nextNPIcell)
    # nextNPI = nextNPIcell.value()
    # print(nextNPI)
    # prevNPIcell = sheet_obj.cell(row = i-1, column = 1)
    # prevNPI = prevNPIcell.value()
    #b = 0
    print("in readExcel, data from excel: ")
    print(Prscrbr_NPI)
    print(Prscrbr_Last_Org_Name)
    print(Prscrbr_First_Name)
    print(Prscrbr_City)
    print(Prscrbr_State_Abrvtn)
    
    #figure out dupes here 
    
    
    try:
        print("data in currentDoctor in readExcel: ")
        print(currentDoctor.NPI)
        print(currentDoctor.lastName)
        print(currentDoctor.firstName)
        print(currentDoctor.city)
        print(currentDoctor.state)
    except: 
        print("no data in currentDoctor")
        pass
    if not currentDoctor: 
        currentDoctor = Doctor(Prscrbr_NPI, Prscrbr_Last_Org_Name, Prscrbr_First_Name, Prscrbr_City, Prscrbr_State_Abrvtn)
        return currentDoctor, i

    else:
        
        currentDoctor = Doctor(Prscrbr_NPI, Prscrbr_Last_Org_Name, Prscrbr_First_Name, Prscrbr_City, Prscrbr_State_Abrvtn)
        return currentDoctor, i
            
    
class Doctor:
    #this is used to save the data while working on each doctor
    def __init__(self, NPI, lastName, firstName, city, state):
        self.lastName = lastName
        self.firstName = firstName
        self.city = city
        self.state = state
        self.clinicName = ''
        self.clinicAddress = ''
        self.clinicPhone = ''
        self.NPI = NPI
    
    def add_clinic_name(self, clinicName):
        self.clinicName = clinicName
    def add_clinic_address(self, clinicAddress):
        self.clinicAddress = clinicAddress
    def add_clinic_phone(self, clinicPhone):
        self.clinicPhone = clinicPhone 
    def read_doctor(self):
        print("info in currentDoctor: ")
        print(self.NPI)
        print(self.lastName)
        print(self.firstName)
        print(self.city)
        print(self.state)
        
        
def main():
    failedDrs = []
    Excels = os.path.join(os.path.dirname(__file__), "Excels")  #put excel files to be processed in a folder named Excels in the same directory as where this program is
    fileList = []
    currentDoctor = False
    for entry in os.scandir(Excels):
        fileList.append(entry.path)
    print(fileList)
    for file in fileList:
        dupesDict = countDupes(file)
        print("opening: " + file)
        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row  #this is the # of rows in the current spreadsheet
        print("M-row is: " + str(m_row))
        driver = webdriver.Chrome()
        i = 43500  #this should start at 2 since 1 is the row where each column is named 
        while i < m_row + 1:
            
            print("in for i in main loop, i: " + str(i))
            
            currentDoctor, i = readExcel(currentDoctor, sheet_obj, i, driver, dupesDict)
             
            print("i after readExcel: " + str(i))
            if findDoctor(driver, currentDoctor):  #if this is true then we are skipping due to doctor not being in healthgrades 
                print("Unable to find doctor in healthgrades!")
                currentDoctor.add_clinic_name("Doctor not found in healthgrades")
                i+=1
                continue
                
                
            time.sleep(7)
            getData(driver, currentDoctor)
            i = logData(i, currentDoctor, sheet_obj, driver, dupesDict)
            # saveCounter = saveCounter + 1
            # if i % 10 == 0:
            #     print("saving for 10 rows")
            i = i+1
            print("saving data")
            wb_obj.save(file)
            
        print("***saving data, do not close***")
        wb_obj.save(file)
        wb_obj.close()
        

if __name__ == '__main__':        
    #driver = webdriver.Chrome()
    main()
    